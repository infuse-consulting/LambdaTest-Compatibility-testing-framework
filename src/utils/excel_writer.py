import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_FILE = os.path.join(BASE_DIR, "reports", "compatibility_test_report.xlsx")
def write_result(result, excel_file=EXCEL_FILE):
    df = pd.DataFrame([result])
    try:
        file_exists = os.path.exists(excel_file)

        if not file_exists:
            df.to_excel(excel_file, index=False)
            book = load_workbook(excel_file)
            sheet = book.active
            current_data_row = 2
        else:
            temp_book = load_workbook(excel_file)
            temp_sheet = temp_book.active
            old_max_row = temp_sheet.max_row

            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, startrow=old_max_row)
            book = load_workbook(excel_file)
            sheet = book.active
            current_data_row = sheet.max_row

        video_link = result.get("Video Link")
        if video_link and sheet:
            col_idx = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == "Video Link":
                    col_idx = col
                    break

            if col_idx:
                cell = sheet.cell(row=current_data_row, column=col_idx)
                cell.value = video_link
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=video_link)
                cell.font = Font(color="0000FF", underline="single")
                cell.alignment = Alignment(horizontal='left')

        book.save(excel_file)
    except PermissionError:
        print(f"Excel file '{excel_file}' is locked. Please close the file and retry.")
    except Exception as e:
        print(f"Failed to write to Excel '{excel_file}': {e}")